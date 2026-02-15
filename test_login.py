import allure
from openpyxl import load_workbook
from playwright.sync_api import Browser, expect


@allure.title("Login valid")
@allure.tag("Login")
def test_login_valid(browser: Browser):
    context = browser.new_context()
    page = context.new_page()
    page.set_viewport_size({"width": 1920, "height": 1080})

    with allure.step("Open browser"):
        page.goto("https://practice.qabrains.com/ecommerce/login")

    with allure.step("Check and validate current url"):
        expect(page).to_have_url("https://practice.qabrains.com/ecommerce/login")

    with allure.step("Check and validate email field"):
        expect(page.get_by_label("email")).to_be_visible()
        expect(page.get_by_label("email")).to_be_enabled()
        page.get_by_label("email").fill("test@qabrains.com")

    with allure.step("Check and validate password field"):
        expect(page.get_by_label("password")).to_be_visible()
        expect(page.get_by_label("password")).to_be_enabled()
        page.get_by_label("password").fill("Password123")

    login_page_bytes = page.screenshot()
    allure.attach(
        login_page_bytes,
        name="login valid page",
        attachment_type=allure.attachment_type.PNG,
    )

    with allure.step("Check and validate login button"):
        expect(page.get_by_role("button", name="Login")).to_be_visible()
        expect(page.get_by_role("button", name="Login")).to_be_enabled()
        page.get_by_role("button", name="Login").click()

    with allure.step("Check and validate current url"):
        expect(page).to_have_url("https://practice.qabrains.com/ecommerce")


@allure.title("Login invalid")
@allure.tag("Login")
def test_login_invalid(browser: Browser):
    context = browser.new_context()
    page = context.new_page()
    page.set_viewport_size({"width": 1920, "height": 1080})

    with allure.step("Open browser"):
        page.goto("https://practice.qabrains.com/ecommerce/login")

    with allure.step("Check and validate current url"):
        expect(page).to_have_url("https://practice.qabrains.com/ecommerce/login")

    with allure.step("All field is empty"):
        with allure.step("Check and validate login button"):
            expect(page.get_by_role("button", name="Login")).to_be_visible()
            expect(page.get_by_role("button", name="Login")).to_be_enabled()
            page.get_by_role("button", name="Login").click()

        with allure.step("Check and validate notification invalid"):
            expect(page.get_by_text("Email is a required field")).to_be_visible()
            expect(page.get_by_text("Password is a required field")).to_be_visible()

        login_page_bytes = page.screenshot()
        allure.attach(
            login_page_bytes,
            name="login invalid (all fields is empty) page",
            attachment_type=allure.attachment_type.PNG,
        )

    with allure.step("Read excel data"):
        wb = load_workbook("qa-brains.xlsx")
        ws = wb["login"]

    with allure.step("Email is not empty and password is empty"):
        page.reload()

        with allure.step("Check and validate email field"):
            expect(page.get_by_label("email")).to_be_visible()
            expect(page.get_by_label("email")).to_be_enabled()
            email = ws["B2"].value
            page.get_by_label("email").fill(email)

        with allure.step("Check and validate login button"):
            expect(page.get_by_role("button", name="Login")).to_be_visible()
            expect(page.get_by_role("button", name="Login")).to_be_enabled()
            page.get_by_role("button", name="Login").click()

        with allure.step("Check and validate notification invalid"):
            expect(page.get_by_text("Password is a required field")).to_be_visible()

        login_page_bytes = page.screenshot()
        allure.attach(
            login_page_bytes,
            name="login invalid (password field is empty) page",
            attachment_type=allure.attachment_type.PNG,
        )

    with allure.step("Email is empty and password is not empty"):
        page.reload()

        with allure.step("Check and validate password field"):
            expect(page.get_by_label("password")).to_be_visible()
            expect(page.get_by_label("password")).to_be_enabled()
            password = ws["C3"].value
            page.get_by_label("password").fill(password)

        with allure.step("Check and validate login button"):
            expect(page.get_by_role("button", name="Login")).to_be_visible()
            expect(page.get_by_role("button", name="Login")).to_be_enabled()
            page.get_by_role("button", name="Login").click()

        with allure.step("Check and validate notification invalid"):
            expect(page.get_by_text("Email is a required field")).to_be_visible()

        login_page_bytes = page.screenshot()
        allure.attach(
            login_page_bytes,
            name="login invalid (email field is empty) page",
            attachment_type=allure.attachment_type.PNG,
        )

    with allure.step("All fields is not empty"):
        page.reload()

        with allure.step("Check and validate email field"):
            expect(page.get_by_label("email")).to_be_visible()
            expect(page.get_by_label("email")).to_be_enabled()
            email = ws["B4"].value
            page.get_by_label("email").fill(email)

        with allure.step("Check and validate password field"):
            expect(page.get_by_label("password")).to_be_visible()
            expect(page.get_by_label("password")).to_be_enabled()
            password = ws["C4"].value
            page.get_by_label("password").fill(password)

        with allure.step("Check and validate login button"):
            expect(page.get_by_role("button", name="Login")).to_be_visible()
            expect(page.get_by_role("button", name="Login")).to_be_enabled()
            page.get_by_role("button", name="Login").click()

        with allure.step("Check and validate notification invalid"):
            expect(page.get_by_text("Username is incorrect.")).to_be_visible()
            expect(page.get_by_text("Password is incorrect.")).to_be_visible()
            expect(
                page.get_by_text("Neither email nor password matched.")
            ).to_be_visible()

        login_page_bytes = page.screenshot()
        allure.attach(
            login_page_bytes,
            name="login invalid (user doesn't exist) page",
            attachment_type=allure.attachment_type.PNG,
        )

    wb.close()
